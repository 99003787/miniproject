# importing libraries for better implementation of code
import pandas as pd
from openpyxl import load_workbook
# read function has been  defined for read all the sheets in  excel file
z = pd.read_excel('python excel final.xlsx', sheet_name=['Sheet1',
                                                         'Sheet2', 'Sheet3',
                                                         'Sheet4', 'Sheet5'])
# Creating empty list
tmp = []
# creating empty data frame
df3 = pd.DataFrame()
# taking input from the user
n = int(input('Enter no of inputs:-'))
count = 0
for _ in range(n):
    tmp1 = []
# taking inputs and append all the data in to empty  list
    h = int(input('Enter your ps no:-'))
    name = str(input('Enter the Name:-'))
    email = str(input('Enter the email:-'))
    tmp1.append(h)
    tmp1.append(name)
    tmp1.append(email)
    tmp.append(tmp1)
# taking all the columns from 5 sheets in to data frame df1
df1 = pd.DataFrame(columns=['SL#', 'PS number', 'Display Name',
                            'Official Email Address', 'company name',
                            'Year of join', 'Room No', 'Block', 'Area',
                            'location', 'Training Room',
                            'Training Name', 'Domain', 'c lang',
                            'linux', 'python', 'Salary',
                            'Designation', 'Blood group', 'Gender',
                            'Phone no', 'Aadhaar Num', 'In Time',
                            'Out Time', 'BUS NUM', 'Attendance',
                            'Self declaration', 'Temperature',
                            'Age', 'Marital status', 'DOB',
                            'Country', 'State', 'Initial'])
# taking input entries and storing in i
for i in tmp:
    h, name, email = i
# searching data in sheets by iterating
    y = z['Sheet1']

    y = y[(y['PS number'] == h) & (y['Display Name'] == name) & (y['Official Email Address'] == email)]
# if the length is  equals to any number then it will move to for loop
# else it prints no match
    if len(y) == 0:
        print('No match')
    else:
        df = pd.DataFrame(y, columns=['SL#', 'PS number', 'Display Name', 'Official Email Address'])
        for i in z.keys():
            x = z[i]
            t = x[(x['PS number'] == h) & (x['Display Name'] == name) & (x['Official Email Address'] == email)]
            col = x.columns
            for j in col:
                df[j] = t[j]
                count = count+1
                df3.at[i, 'H'] = count
    df1 = df1.append(df)
df2 = df1.describe()
df3.at[1, 'Total columns'] = (len(df1.columns)*n)
# load_workbook( ) function is used
# when you have to access an MS Excel file in openpyxl module.
# load workbook function only works if
# you have an already created file on your disk
# and you want to open workbook for some operation.
book = load_workbook('python excel final.xlsx')
writer = pd.ExcelWriter('python excel final.xlsx', engine='openpyxl')
writer.book = book
# ExcelWriter for some reason uses
# writer.sheets to access the sheet.
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
# writing the data into the master sheet
df1.to_excel(writer, sheet_name='master', index=False)
# writing the data into the summary sheet
df3.to_excel(writer, sheet_name='summary', index=False)
# saving  the data in the master sheet
writer.save()
